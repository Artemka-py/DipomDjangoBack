from django.contrib import admin
from django.contrib.auth.models import Group
from rangefilter.filter import DateRangeFilter
from import_export import resources
from import_export.admin import ImportExportModelAdmin
from . import models
from django.http import HttpResponse
from datetime import datetime
from mptt.admin import DraggableMPTTAdmin
from unidecode import unidecode
import csv, datetime
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin

from .forms import UserCreationForm


def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment
    from datetime import datetime
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={}-{}.xlsx'.format(
        unidecode(modeladmin.model._meta.verbose_name), datetime.now().strftime('%Y-%m-%d'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = unidecode(modeladmin.model._meta.verbose_name)
    row_num = 0

    columns = [
        (u"Название задачи", 30),
        (u"Название этапа", 30),
        (u"Логин разработчика", 30),
        (u"Название главной задачи", 30),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.font = Font(bold=True)
        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.task_name,
            str(obj.task_stage),
            str(obj.task_developer_login),
            str(obj.parent),
        ]
        for col_num in range(len(row)):
            if (obj.parent):
                c = ws.cell(row=row_num + 1, column=col_num + 2)
            else:
                c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            if (obj.parent):
                c.font = Font(color="00FF6600")
            c.alignment = Alignment(wrapText=True)

    wb.save(response)
    return response


export_xlsx.short_description = "Экспорт в excel"


def export_to_csv(modeladmin, request, queryset):
    opts = modeladmin.model._meta
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename={}.csv'.format(
        unidecode(modeladmin.model._meta.verbose_name))
    writer = csv.writer(response)
    fields = [field for field in opts.get_fields() if not field.many_to_many and not field.one_to_many]
    writer.writerow([field.verbose_name for field in fields])
    for obj in queryset:
        data_row = []
        for field in fields:
            value = getattr(obj, field.name)
            if isinstance(value, datetime.datetime):
                value = value.strftime('%d/%m/%Y')
            data_row.append(value)
        writer.writerow(data_row)

    return response


export_to_csv.short_description = 'Экспорт в CSV'


def log_del_True(modeladmin, request, queryset):
    queryset.filter(logical_delete_status=False).update(logical_delete_status=True)


log_del_True.short_description = 'Поменять статус логического удаления на правда'


def log_del_False(modeladmin, request, queryset):
    queryset.filter(logical_delete_status=True).update(logical_delete_status=False)


log_del_False.short_description = 'Поменять статус логического удаления на ложь'


class WorkGroupsResource(resources.ModelResource):
    class Meta:
        model = models.Workgroups
        import_id_fields = ('workgroup_id',)
        exclude = ('workgroup_id',)
        # fields = ('workgroup_name')

    def before_import(self, dataset, dry_run):
        if dataset.headers:
            dataset.headers = [str(header).lower().strip() for header in dataset.headers]


class UserResource(resources.ModelResource):
    class Meta:
        model = models.Users
        # exclude = 'user_password'
        import_id_fields = ['username']


# Классы для кастомизации моделей на сайте в админке
class UsersModel(BaseUserAdmin, ImportExportModelAdmin):
    add_form = UserCreationForm
    resource_class = UserResource
    # fields = (('username', 'password'), ('first_name', 'middle_name', 'sur_name'), 'birth_date', 'phone_num',
    #           'email', 'is_Active', 'user_image_src')
    fieldsets = (
        (None, {'fields': ('username', 'password', 'email')}),
        ('Персональные данные:', {'fields': ('first_name', 'middle_name', 'sur_name',
                                             'birth_date', 'phone_num', 'user_image_src')}),
        ('Доступы', {'fields': ('is_admin', 'is_staff', 'is_Active')})
    )

    list_display = ('username', 'first_name', 'middle_name', 'sur_name', 'birth_date', 'image_img',
                    'phone_num', 'email', 'is_Active')
    list_per_page = 10
    search_fields = (
        'username', 'first_name', 'middle_name', 'sur_name', 'birth_date', 'phone_num',
        'email'
    )
    list_filter = ('first_name', 'middle_name', 'sur_name',
                   ('birth_date', DateRangeFilter), 'is_Active')
    actions = [log_del_True, log_del_False]

    filter_horizontal = ()

    class Media:
        pass


class TaskModel(DraggableMPTTAdmin):
    list_display = (
        'tree_actions',
        'indented_title',
        'task_developer_login'
    )
    actions = [export_to_csv, export_xlsx]


class WorkgroupsModel(ImportExportModelAdmin):
    resources = WorkGroupsResource


@admin.register(models.Status)
class Test(admin.ModelAdmin):
    list_display = ['status_name']


@admin.register(models.Stages)
class StagesAdmin(admin.ModelAdmin):
    exclude = ['stage_id']


@admin.register(models.Projects)
class ProjectsAdmin(admin.ModelAdmin):
    list_display = ('project_name', 'project_status', 'start_date_plan', 'finish_date_plan', 'tech_task_path')


# Register your models here.
admin.site.register(models.Users, UsersModel)
admin.site.register(models.Tasks, TaskModel)
admin.site.register(models.Managers, list_display=('manager_login', 'outsource_spec',))
admin.site.register(models.Clients, list_display=('client_login', 'client_organisation',))
admin.site.register(models.Modules)
admin.site.register(models.Developers)
admin.site.register(models.Workgroups, WorkgroupsModel)
admin.site.register(models.Organisations, list_display=('full_name', 'short_name', 'image_img',))
admin.site.register(models.WorkingDeveloperList)
admin.site.register(models.Issues)
admin.site.register(models.Notes)
admin.site.unregister(Group)
